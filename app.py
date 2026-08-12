import streamlit as st
import pandas as pd
import plotly.express as px
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
import gc

# -----------------------------------------------------------------------------
# CONFIGURAÇÃO DA PÁGINA
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title="Processador Operacional - Leituras e Entregas",
    page_icon="⚡",
    layout="wide"
)

st.title("⚡ Processador Operacional (Leituras & Entregas)")
st.markdown("Cruzamento inteligente de Lotes e tratamento otimizado contra estouro de memória RAM.")
st.markdown("---")

# -----------------------------------------------------------------------------
# FUNÇÃO AUXILIAR DE LEITURA EM CHUNKS (EVITA CRASH)
# -----------------------------------------------------------------------------
def ler_arquivo_em_chunks(file, chunk_size=50000):
    """Lê o arquivo enviado em lotes sem carregar 200MB de uma só vez na memória."""
    file.seek(0)
    chunks = []
    
    # Se for CSV
    if file.name.lower().endswith('.csv'):
        # Tenta ler com engine C e separador automático
        try:
            for chunk in pd.read_csv(file, chunksize=chunk_size, low_memory=False, on_bad_lines='skip', sep=None, engine='python'):
                # Otimiza memória convertendo 'object' para 'category'
                for col in chunk.select_dtypes(include='object').columns:
                    chunk[col] = chunk[col].astype('category')
                chunks.append(chunk)
        except Exception:
            file.seek(0)
            # Fallback para latin1 caso encoding padrão falhe
            for chunk in pd.read_csv(file, chunksize=chunk_size, low_memory=False, encoding='latin1', on_bad_lines='skip', sep=';'):
                for col in chunk.select_dtypes(include='object').columns:
                    chunk[col] = chunk[col].astype('category')
                chunks.append(chunk)
        
        if chunks:
            df = pd.concat(chunks, ignore_index=True)
            del chunks
            gc.collect()
            return df
            
    # Se for Excel (.xlsx)
    else:
        df = pd.read_excel(file)
        for col in df.select_dtypes(include='object').columns:
            df[col] = df[col].astype('category')
        return df
        
    return None

# -----------------------------------------------------------------------------
# FUNÇÃO PRINCIPAL DE PROCESSAMENTO
# -----------------------------------------------------------------------------
@st.cache_data(show_spinner=False, max_entries=2)
def processar_arquivos_unificados(uploaded_files):
    dfs_leitura = []
    dfs_entrega = []

    for file in uploaded_files:
        df_temp = ler_arquivo_em_chunks(file)

        if df_temp is None or df_temp.empty:
            continue

        # Normaliza nomes das colunas
        df_temp.columns = df_temp.columns.astype(str).str.strip().str.upper()

        # Limpeza de caracteres especiais leve
        for col in df_temp.select_dtypes(include=['category', 'object']).columns:
            df_temp[col] = (
                df_temp[col]
                .astype(str)
                .str.replace(r'[\r\n]+', ' ', regex=True)
                .str.replace('"', '')
                .str.strip()
            )

        is_entrega = ('DATA_HORA_APROXIMADA' in df_temp.columns or 'DAT_PREVISTA_ENTREGA' in df_temp.columns)

        if is_entrega:
            dfs_entrega.append(df_temp)
        else:
            dfs_leitura.append(df_temp)

    mapa_unidade_lote = {}
    dfs_processados = []

    # 1. Processamento de Leituras
    if dfs_leitura:
        df_leitura_concat = pd.concat(dfs_leitura, ignore_index=True)
        del dfs_leitura
        gc.collect()

        col_dt = 'DT_INI_ACAO' if 'DT_INI_ACAO' in df_leitura_concat.columns else 'DAT_PREVISTA'
        col_base = 'NOM_BASE_OPERACIONAL' if 'NOM_BASE_OPERACIONAL' in df_leitura_concat.columns else 'BASE_OPERACIONAL'
        col_mun = 'NOM_MUNICIPIO' if 'NOM_MUNICIPIO' in df_leitura_concat.columns else 'MUNICIPIO'
        col_lote = 'LOTE' if 'LOTE' in df_leitura_concat.columns else 'NUM_LOTE'
        col_unidade = 'NOM_UNIDADE_LEITURA' if 'NOM_UNIDADE_LEITURA' in df_leitura_concat.columns else 'UNIDADE_LEITURA'
        col_cod_agente = 'COD_AGENTE' if 'COD_AGENTE' in df_leitura_concat.columns else 'CODIGO_AGENTE'
        col_nom_agente = 'AGENTE' if 'AGENTE' in df_leitura_concat.columns else 'NOM_AGENTE'
        col_loc = 'LOCALIZACAO' if 'LOCALIZACAO' in df_leitura_concat.columns else 'ZONA'
        col_ind_tipo = 'IND_TIPO' if 'IND_TIPO' in df_leitura_concat.columns else 'TIPO'
        col_tipo_atv = 'TIPO_ATIVIDADE' if 'TIPO_ATIVIDADE' in df_leitura_concat.columns else 'TIPO_SERVICO'
        col_nota = 'COD_NOTA_VISITA' if 'COD_NOTA_VISITA' in df_leitura_concat.columns else None

        if col_ind_tipo in df_leitura_concat.columns:
            df_leitura_concat = df_leitura_concat[df_leitura_concat[col_ind_tipo].astype(str).str.strip().isin(['P', 'R', 'C'])].copy()

        if not df_leitura_concat.empty:
            df_padrao_leituras = pd.DataFrame()

            dt_series = pd.to_datetime(df_leitura_concat[col_dt], dayfirst=True, errors='coerce') if col_dt in df_leitura_concat.columns else pd.Series(pd.NaT, index=df_leitura_concat.index)

            df_padrao_leituras['DATA_HORA_DT'] = dt_series
            df_padrao_leituras['DATA_REAL'] = dt_series.dt.strftime('%d/%m/%Y').fillna('Sem Data')
            df_padrao_leituras['HORA'] = dt_series.dt.strftime('%H:%M').fillna('N/A')

            df_padrao_leituras['BASE_STD'] = df_leitura_concat[col_base].fillna('Não Informado').replace({'': 'Não Informado', 'nan': 'Não Informado'}).astype(str).str.strip() if col_base in df_leitura_concat.columns else 'Não Informado'
            df_padrao_leituras['MUNICIPIO_STD'] = df_leitura_concat[col_mun].fillna('Não Informado').replace({'': 'Não Informado', 'nan': 'Não Informado'}).astype(str).str.strip() if col_mun in df_leitura_concat.columns else 'Não Informado'

            lote_s = df_leitura_concat[col_lote].fillna('').astype(str).str.strip() if col_lote in df_leitura_concat.columns else pd.Series('', index=df_leitura_concat.index)
            df_padrao_leituras['LOTE_STD'] = lote_s.replace({'': '(Sem Lote)', 'nan': '(Sem Lote)'})

            unidade_s = df_leitura_concat[col_unidade].fillna('Não Informado').astype(str).str.strip() if col_unidade in df_leitura_concat.columns else pd.Series('Não Informado', index=df_leitura_concat.index)
            df_padrao_leituras['UNIDADE_STD'] = unidade_s.replace({'': 'Não Informado', 'nan': 'Não Informado'})

            # Criação do Mapa de Lotes
            valido_map = df_padrao_leituras[(df_padrao_leituras['UNIDADE_STD'] != 'Não Informado') & (df_padrao_leituras['LOTE_STD'] != '(Sem Lote)')]
            if not valido_map.empty:
                mapa_unidade_lote = valido_map.groupby('UNIDADE_STD')['LOTE_STD'].first().to_dict()

            cod_ag = df_leitura_concat[col_cod_agente].fillna('Sem Código').astype(str).str.strip().str.replace(r'\.0$', '', regex=True) if col_cod_agente in df_leitura_concat.columns else pd.Series('Sem Código', index=df_leitura_concat.index)
            df_padrao_leituras['COD_AGENTE_STD'] = cod_ag.replace({'nan': 'Sem Código', '': 'Sem Código'})

            nom_ag = df_leitura_concat[col_nom_agente].fillna('').astype(str).str.strip() if col_nom_agente in df_leitura_concat.columns else pd.Series('', index=df_leitura_concat.index)
            df_padrao_leituras['NOM_AGENTE_STD'] = nom_ag.replace({'nan': ''})

            loc_s = df_leitura_concat[col_loc].fillna('').astype(str).str.strip() if col_loc in df_leitura_concat.columns else pd.Series('', index=df_leitura_concat.index)
            df_padrao_leituras['LOCALIZACAO_STD'] = loc_s.replace({'': 'Não Informado', 'nan': 'Não Informado'})

            df_padrao_leituras['IND_TIPO_STD'] = df_leitura_concat[col_ind_tipo].astype(str).str.strip()

            atv_s = df_leitura_concat[col_tipo_atv].fillna('Leitura').astype(str).str.strip() if col_tipo_atv in df_leitura_concat.columns else pd.Series('Leitura', index=df_leitura_concat.index)
            df_padrao_leituras['TIPO_ATIVIDADE_STD'] = atv_s.replace({'': 'Leitura', 'nan': 'Leitura'})

            df_padrao_leituras['TAREFA_STD'] = df_leitura_concat.index

            if col_nota and col_nota in df_leitura_concat.columns:
                nota_series = df_leitura_concat[col_nota].fillna('').astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
            else:
                nota_series = pd.Series('', index=df_leitura_concat.index)

            df_padrao_leituras['IMP_GRUPO_1'] = nota_series.str.startswith('1').astype(int)
            df_padrao_leituras['IMP_GRUPO_2'] = nota_series.str.startswith('2').astype(int)
            df_padrao_leituras['TOTAL_IMP'] = df_padrao_leituras['IMP_GRUPO_1'] + df_padrao_leituras['IMP_GRUPO_2']
            df_padrao_leituras['LEITURA_LIMPA'] = (df_padrao_leituras['TOTAL_IMP'] == 0).astype(int)
            df_padrao_leituras['ORIGEM_DADO'] = "Leitura"

            dfs_processados.append(df_padrao_leituras)
            del df_leitura_concat
            gc.collect()

    # 2. Processamento de Entregas
    if dfs_entrega:
        df_entrega_concat = pd.concat(dfs_entrega, ignore_index=True)
        del dfs_entrega
        gc.collect()

        col_dt = 'DATA_HORA_APROXIMADA' if 'DATA_HORA_APROXIMADA' in df_entrega_concat.columns else 'DAT_PREVISTA_ENTREGA'
        col_base = 'NOM_BASE_OPERACIONAL' if 'NOM_BASE_OPERACIONAL' in df_entrega_concat.columns else 'BASE_OPERACIONAL'
        col_mun = 'NOM_MUNICIPIO' if 'NOM_MUNICIPIO' in df_entrega_concat.columns else 'MUNICIPIO'
        col_unidade = 'NOM_UNIDADE_LEITURA' if 'NOM_UNIDADE_LEITURA' in df_entrega_concat.columns else 'UNIDADE_LEITURA'
        col_agente = 'COD_AGENTE_COMERCIAL' if 'COD_AGENTE_COMERCIAL' in df_entrega_concat.columns else 'COD_AGENTE'
        col_tarefa = 'SEQ_TAREFA' if 'SEQ_TAREFA' in df_entrega_concat.columns else df_entrega_concat.index

        df_padrao_entregas = pd.DataFrame()
        dt_series = pd.to_datetime(df_entrega_concat[col_dt], dayfirst=True, errors='coerce') if col_dt in df_entrega_concat.columns else pd.Series(pd.NaT, index=df_entrega_concat.index)

        df_padrao_entregas['DATA_HORA_DT'] = dt_series
        df_padrao_entregas['DATA_REAL'] = dt_series.dt.strftime('%d/%m/%Y').fillna('Sem Data')
        df_padrao_entregas['HORA'] = dt_series.dt.strftime('%H:%M').fillna('N/A')

        df_padrao_entregas['BASE_STD'] = df_entrega_concat[col_base].fillna('Não Informado').replace({'': 'Não Informado', 'nan': 'Não Informado'}).astype(str).str.strip() if col_base in df_entrega_concat.columns else 'Não Informado'
        df_padrao_entregas['MUNICIPIO_STD'] = df_entrega_concat[col_mun].fillna('Não Informado').replace({'': 'Não Informado', 'nan': 'Não Informado'}).astype(str).str.strip() if col_mun in df_entrega_concat.columns else 'Não Informado'

        unidade_e = df_entrega_concat[col_unidade].fillna('Não Informado').astype(str).str.strip() if col_unidade in df_entrega_concat.columns else pd.Series('Não Informado', index=df_entrega_concat.index)
        df_padrao_entregas['UNIDADE_STD'] = unidade_e.replace({'': 'Não Informado', 'nan': 'Não Informado'})

        # Cruzamento rápido usando .map (muito mais rápido que .apply)
        df_padrao_entregas['LOTE_STD'] = df_padrao_entregas['UNIDADE_STD'].map(mapa_unidade_lote).fillna("(Sem Lote Cruzado)")

        cod_ag = df_entrega_concat[col_agente].fillna('Sem Código').astype(str).str.strip().str.replace(r'\.0$', '', regex=True) if col_agente in df_entrega_concat.columns else pd.Series('Sem Código', index=df_entrega_concat.index)
        df_padrao_entregas['COD_AGENTE_STD'] = cod_ag.replace({'nan': 'Sem Código', '': 'Sem Código'})
        df_padrao_entregas['NOM_AGENTE_STD'] = ""

        df_padrao_entregas['LOCALIZACAO_STD'] = "(N/A - Entrega)"
        df_padrao_entregas['IND_TIPO_STD'] = "E"
        df_padrao_entregas['TIPO_ATIVIDADE_STD'] = "Entrega"
        df_padrao_entregas['TAREFA_STD'] = df_entrega_concat[col_tarefa] if col_tarefa in df_entrega_concat.columns else df_entrega_concat.index

        df_padrao_entregas['IMP_GRUPO_1'] = 0
        df_padrao_entregas['IMP_GRUPO_2'] = 0
        df_padrao_entregas['TOTAL_IMP'] = 0
        df_padrao_entregas['LEITURA_LIMPA'] = 1
        df_padrao_entregas['ORIGEM_DADO'] = "Entrega"

        dfs_processados.append(df_padrao_entregas)
        del df_entrega_concat
        gc.collect()

    if not dfs_processados:
        return None

    df_concat = pd.concat(dfs_processados, ignore_index=True)
    del dfs_processados
    gc.collect()

    # Mapeamento do Agente Completo
    agentes_map = df_concat[df_concat['NOM_AGENTE_STD'] != ''].groupby('COD_AGENTE_STD')['NOM_AGENTE_STD'].first().to_dict()

    df_concat['NOM_MAPPED'] = df_concat['COD_AGENTE_STD'].map(agentes_map).fillna('')
    df_concat['NOM_FINAL'] = df_concat['NOM_AGENTE_STD'].where(df_concat['NOM_AGENTE_STD'] != '', df_concat['NOM_MAPPED'])
    
    df_concat['AGENTE_COMPLETO'] = df_concat.apply(
        lambda r: f"{r['COD_AGENTE_STD']} - {r['NOM_FINAL']}" if r['NOM_FINAL'] else r['COD_AGENTE_STD'], axis=1
    )
    
    df_concat.drop(columns=['NOM_MAPPED', 'NOM_FINAL'], inplace=True)

    return df_concat

# -----------------------------------------------------------------------------
# INTERFACE PRINCIPAL
# -----------------------------------------------------------------------------
uploaded_files = st.file_uploader(
    "📁 Envie suas planilhas de Leitura e/ou Entrega (.csv ou .xlsx):",
    type=["csv", "xlsx"],
    accept_multiple_files=True
)

if uploaded_files:
    with st.spinner("🚀 Processando arquivos em lotes..."):
        df = processar_arquivos_unificados(uploaded_files)

    if df is not None and not df.empty:
        st.sidebar.header("🎯 Filtros Unificados")

        def criar_multiselect(label, col_name):
            opcoes = sorted([str(x) for x in df[col_name].unique() if str(x) not in ['nan', 'nan - nan']])
            return st.sidebar.multiselect(label, options=opcoes, default=opcoes)

        f_origem = criar_multiselect("Origem do Dado", 'ORIGEM_DADO')
        f_base = criar_multiselect("Base Operacional", 'BASE_STD')
        f_mun = criar_multiselect("Município", 'MUNICIPIO_STD')
        f_unidade = criar_multiselect("Unidade de Leitura", 'UNIDADE_STD')
        f_lote = criar_multiselect("Lote (Cruzado)", 'LOTE_STD')
        f_loc = criar_multiselect("Localização", 'LOCALIZACAO_STD')
        f_ind_tipo = criar_multiselect("IND_TIPO (E, P, R, C)", 'IND_TIPO_STD')
        f_tipo_atv = criar_multiselect("Tipo de Atividade", 'TIPO_ATIVIDADE_STD')
        f_agente = criar_multiselect("Agente Comercial", 'AGENTE_COMPLETO')
        f_data_real = criar_multiselect("Data da Ação", 'DATA_REAL')

        df_filtrado = df[
            (df['ORIGEM_DADO'].astype(str).isin(f_origem)) &
            (df['BASE_STD'].astype(str).isin(f_base)) &
            (df['MUNICIPIO_STD'].astype(str).isin(f_mun)) &
            (df['UNIDADE_STD'].astype(str).isin(f_unidade)) &
            (df['LOTE_STD'].astype(str).isin(f_lote)) &
            (df['LOCALIZACAO_STD'].astype(str).isin(f_loc)) &
            (df['IND_TIPO_STD'].astype(str).isin(f_ind_tipo)) &
            (df['TIPO_ATIVIDADE_STD'].astype(str).isin(f_tipo_atv)) &
            (df['AGENTE_COMPLETO'].astype(str).isin(f_agente)) &
            (df['DATA_REAL'].astype(str).isin(f_data_real))
        ]

        if df_filtrado.empty:
            st.warning("⚠️ Nenhum registro encontrado para a combinação de filtros selecionada.")
        else:
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total Registros", f"{len(df_filtrado):,}".replace(",", "."))
            col2.metric("Registros Limpos", f"{df_filtrado['LEITURA_LIMPA'].sum():,}".replace(",", "."))
            col3.metric("Impedimentos G1", f"{df_filtrado['IMP_GRUPO_1'].sum():,}".replace(",", "."))
            col4.metric("Impedimentos G2", f"{df_filtrado['IMP_GRUPO_2'].sum():,}".replace(",", "."))

            total_geral_imp = df_filtrado['TOTAL_IMP'].sum()
            st.info(f"🚨 **Total Geral de Impedimentos Operacionais:** {total_geral_imp:,}".replace(",", "."))

            st.markdown("---")

            def hora_min_max(s, tipo):
                v = s.dropna()
                if v.empty: return "N/A"
                res = v.min() if tipo == 'min' else v.max()
                return res.strftime('%H:%M') if pd.notna(res) else "N/A"

            df_resumo = df_filtrado.groupby([
                'DATA_REAL', 'BASE_STD', 'MUNICIPIO_STD', 'LOTE_STD', 'UNIDADE_STD',
                'LOCALIZACAO_STD', 'IND_TIPO_STD', 'TIPO_ATIVIDADE_STD', 'AGENTE_COMPLETO'
            ], as_index=False).agg(
                TOTAL_REGISTROS=('TAREFA_STD', 'count'),
                LEITURAS_LIMPAS=('LEITURA_LIMPA', 'sum'),
                IMP_G1=('IMP_GRUPO_1', 'sum'),
                IMP_G2=('IMP_GRUPO_2', 'sum'),
                TOTAL_IMP=('TOTAL_IMP', 'sum'),
                HORA_INI=('DATA_HORA_DT', lambda x: hora_min_max(x, 'min')),
                HORA_FIM=('DATA_HORA_DT', lambda x: hora_min_max(x, 'max'))
            )

            df_resumo.columns = [
                'Data Realização', 'Base Operacional', 'Município', 'Lote',
                'Unidade de Leitura', 'Localização', 'IND_TIPO', 'Tipo Atividade',
                'Agente Comercial', 'Total Registros', 'Limpos / Sucesso',
                'Impedimentos G1', 'Impedimentos G2', 'Total Impedimentos', '1ª Ação', 'Última Ação'
            ]

            col_graf1, col_graf2 = st.columns(2)
            with col_graf1:
                st.subheader("🏙️ Volume por Município")
                df_cidade = df_filtrado.groupby('MUNICIPIO_STD').size().reset_index(name='Qtd Registros')
                fig_cidade = px.bar(
                    df_cidade, x='MUNICIPIO_STD', y='Qtd Registros', text_auto=True,
                    labels={'MUNICIPIO_STD': 'Município', 'Qtd Registros': 'Volume'},
                    color_discrete_sequence=['#1f77b4']
                )
                fig_cidade.update_layout(xaxis_title="", yaxis_title="")
                st.plotly_chart(fig_cidade, use_container_width=True)

            with col_graf2:
                st.subheader("📊 Produção por IND_TIPO")
                df_ind_graf = df_filtrado.groupby('IND_TIPO_STD').size().reset_index(name='Qtd Registros')
                fig_ind = px.pie(
                    df_ind_graf, names='IND_TIPO_STD', values='Qtd Registros', hole=0.45,
                    color_discrete_sequence=px.colors.qualitative.Set2
                )
                st.plotly_chart(fig_ind, use_container_width=True)

            st.markdown("---")
            st.subheader("📋 Resumo Consolidado por Atividade e Agente")
            st.dataframe(df_resumo, use_container_width=True)

            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df_resumo.to_excel(writer, sheet_name="Resumo Consolidado", index=False)

                df_detalhado_export = df_filtrado[[
                    'ORIGEM_DADO', 'DATA_REAL', 'BASE_STD', 'MUNICIPIO_STD', 'LOTE_STD', 'UNIDADE_STD',
                    'LOCALIZACAO_STD', 'IND_TIPO_STD', 'TIPO_ATIVIDADE_STD', 'AGENTE_COMPLETO', 'HORA',
                    'LEITURA_LIMPA', 'IMP_GRUPO_1', 'IMP_GRUPO_2', 'TOTAL_IMP'
                ]].copy()

                df_detalhado_export.columns = [
                    'Origem', 'Data Realização', 'Base Operacional', 'Município', 'Lote', 'Unidade de Leitura',
                    'Localização', 'IND_TIPO', 'Tipo Atividade', 'Agente Comercial', 'Hora',
                    'Registro Limpo (1/0)', 'Impedimento G1', 'Impedimento G2', 'Total Impedimentos'
                ]
                df_detalhado_export.to_excel(writer, sheet_name="Base Filtrada Detalhada", index=False)

                workbook = writer.book
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

                for ws in workbook.worksheets:
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    for col in ws.columns:
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = 20

            st.download_button(
                label="📥 Baixar Relatório Unificado (Excel)",
                data=buffer.getvalue(),
                file_name="relatorio_unificado_operacional.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    else:
        st.error("Não foi possível extrair dados válidos dos arquivos fornecidos.")
else:
    st.info("👆 Selecione um ou mais arquivos de Leitura e/ou Entrega acima para iniciar o processamento.")
